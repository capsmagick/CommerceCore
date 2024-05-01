import json

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from rest_framework.decorators import action
from django.http import HttpResponse
import datetime


class ExportData:
    EXCLUDE_FIELDS = ['deleted', 'deleted_at', 'deleted_by']
    EMPTY_VALUES = [None, '', ' ', 'undefined', '[]']
    INCLUDE_DELETED = False

    def generate_headers(self, model):
        headers = []
        for field in model._meta.fields:
            if self.INCLUDE_DELETED or field.name not in self.EXCLUDE_FIELDS:
                if not self.SELECTED_FIELDS or field.name in self.SELECTED_FIELDS:
                    headers.append(field.verbose_name)

        # Include many-to-many field headers
        for field in model._meta.many_to_many:
            if field.many_to_many:
                if not self.SELECTED_FIELDS or field.name in self.SELECTED_FIELDS:
                    headers.append(field.verbose_name)

        return headers

    def get_field_data(self, obj, field):
        field_type = field.get_internal_type()
        if self.INCLUDE_DELETED or field.name not in self.EXCLUDE_FIELDS:
            data = getattr(obj, field.name)
            if field_type == 'DateTimeField':
                data = data.strftime("%Y-%m-%d %I:%M %p") if data not in self.EMPTY_VALUES else ''
            elif field_type == 'DateField':
                data = data.strftime("%Y-%m-%d") if data not in self.EMPTY_VALUES else ''
            elif field_type == 'ForeignKey':
                data = str(data) if data not in self.EMPTY_VALUES else ''
            elif field_type == 'BooleanField':
                data = 'Yes' if data else 'No'
            else:
                data = str(data) if data not in self.EMPTY_VALUES else ''

            return data

    @action(detail=False, methods=['GET'], url_path='export')
    def generate_excel(self, request, *args, **kwargs):
        selected_fields = request.GET.get('fields', '[]')
        include_deleted = request.GET.get('includeDeleted', False)

        self.SELECTED_FIELDS = json.loads(selected_fields) # noqa
        self.INCLUDE_DELETED = include_deleted

        queryset = self.filter_queryset(self.get_queryset())
        model = queryset.model

        today = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M")
        file_name = f"{model.__name__}_{today}"

        wb = Workbook()
        ws = wb.active

        headers = self.generate_headers(model)

        for col, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=value)
            cell.font = Font(bold=True)
            cell.border = Border(
                top=Side(style='thin'), bottom=Side(style='thin'),
                right=Side(style='thin'), left=Side(style='thin'),
            )
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = len(value) + 2
            ws.row_dimensions[1].height = 30
            ws.freeze_panes = 'A2'

        for obj in queryset:
            row = []
            for field in model._meta.fields:
                data = self.get_field_data(obj, field)
                if data:
                    if not self.SELECTED_FIELDS or field.name in self.SELECTED_FIELDS:
                        row.append(data)

            # Include many-to-many field data
            for relation in model._meta.many_to_many:
                if relation.many_to_many:
                    related_data = getattr(obj, relation.name).all()
                    data = ', '.join(str(item) for item in related_data) if related_data else ''
                    if not self.SELECTED_FIELDS or relation.name in self.SELECTED_FIELDS:
                        row.append(data)

            ws.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'

        # Save the workbook to the response
        wb.save(response)

        return response





